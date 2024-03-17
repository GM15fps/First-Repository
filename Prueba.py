import streamlit as st
from io import BytesIO
import openpyxl
from tempfile import NamedTemporaryFile

@st.cache_data
def load():
    plantilla = openpyxl.load_workbook('ASTM COTIZACION FER-2024-0118.xlsx')
    plan_hoja = plantilla['Cotizacion']
    return plantilla, plan_hoja
plantilla, plan_hoja = load()

st.title('App')

COTI = st.text_input(label = 'COTI')
COMPONENTE = st.text_input(label = 'COMPONENTE')
OT = st.number_input(label = 'OT')
MS = st.text_input(label = 'MS')
FECHA = st.date_input(label = 'FECHA')
MONTO = st.number_input(label = 'MONTO')

texto = f'''{COMPONENTE}
Comprende :
-Recuperar 01 alojamiento roscado de 3/8 x 16 DP 18 ( TIENE INSERTO).

N/P : {MS}'''

plan_hoja.cell(row = 3, column = 9).value = COTI
plan_hoja.cell(row = 14, column = 4).value = texto
plan_hoja.cell(row = 8, column = 10).value = OT
plan_hoja.cell(row = 9, column = 4).value = FECHA 
plan_hoja.cell(row = 14, column = 10).value = MONTO

with NamedTemporaryFile() as tmp:
     plantilla.save(tmp.name)
     data = BytesIO(tmp.read())

st.download_button("Retrieve file",
     data=data,
     mime='xlsx',
     file_name="Modificado.xlsx")

